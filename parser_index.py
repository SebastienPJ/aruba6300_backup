import os
import parser
import create_sheets
# import object
import pandas as pd
from openpyxl import load_workbook


def parseConfigFile(filepath):

    unfilteredArray = parser.createUnfilteredArray(filepath)
    hostname = parser.getHostname(unfilteredArray)[:-1]

    sheetPath = f'ExcelSheetOutputs/{hostname}.xlsx'

    """""SFP"""
    SfpBeginPoint = ['Port', 'Type', 'Product', 'Serial', 'Part']
    SfpEndPoint =  f'{hostname}#'
    SfpLinesUntilActualDataStart = 3 

    """Members"""
    membersBeginPoint = ['Line', 'Modules']
    membersEndPoint = f'{hostname}#'
    membersLinesUntilActualDataStart = 5

    """Power-Supply"""
    powerSupplyBeginPoint = ['show', 'environment', 'power-supply']
    
    powerSupplyEndPoint = ['show', 'environment', 'power-redundancy']

    if powerSupplyEndPoint not in unfilteredArray:
        powerSupplyEndPoint = ['show', 'environment', 'temperature']

    powerSupplyLinesUntilActualDataStart = 4



    sfpData = create_sheets.createSfpSheet(unfilteredArray, SfpBeginPoint, SfpEndPoint, SfpLinesUntilActualDataStart, sheetPath)  

    membersData = create_sheets.createMembersSheet(unfilteredArray, membersBeginPoint, membersEndPoint, membersLinesUntilActualDataStart, sheetPath)

    powerSupplyData = create_sheets.createPowerSupplySheet(unfilteredArray, powerSupplyBeginPoint, powerSupplyEndPoint, powerSupplyLinesUntilActualDataStart, sheetPath)

    # os.system(f'start EXCEL.EXE {sheetName}')
    # switchStackObj = object.generateStackObject(sfpData, membersData, powerSupplyData)
    # print(switchStackObj['member1']['device'])
    # print(switchStackObj)



    # if 'StackObj' in book.sheetnames:
    #     book.remove(book['StackObj'])


    # df = pd.DataFrame(switchStackObj)
    # df.to_excel(writer, sheet_name='StackObj')

    # print(sfpData)


