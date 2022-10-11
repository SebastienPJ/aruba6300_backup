import os
import os.path
import pandas as pd
import parser
from openpyxl import load_workbook



def createSfpSheet(rawDataArray, startPoint, endPoint, linesUntilDataStart, fileSavePath):

    desiredArray = parser.createDesiredArray(rawDataArray, startPoint, endPoint)

    if desiredArray:
        dataOnlyArray = parser.createDataOnlyArray(desiredArray, linesUntilDataStart)
    else:
        dataOnlyArray = ['No SFPs in Stack']
    
    writer = pd.ExcelWriter(fileSavePath, engine='xlsxwriter')

    df = pd.DataFrame(dataOnlyArray)
    df.to_excel(writer, sheet_name='SFPs', index=False, header=False)
    writer.save()

    return dataOnlyArray



def createMembersSheet(rawDataArray, startPoint, endPoint, linesUntilDataStart, fileSavePath):


    desiredArray = parser.createDesiredArray(rawDataArray, startPoint, endPoint)    
    dataOnlyArray = parser.createDataOnlyArray(desiredArray, linesUntilDataStart)

    book = load_workbook(fileSavePath)        
    writer = pd.ExcelWriter(fileSavePath, engine='openpyxl')
    writer.book = book

    if 'Modules' in book.sheetnames:
        book.remove(book['Modules'])


    df = pd.DataFrame(dataOnlyArray)
    df.to_excel(writer, sheet_name='Modules', index=False, header=False)
    writer.save()

    return dataOnlyArray



def createPowerSupplySheet(rawDataArray, startPoint, endPoint, linesUntilDataStart, fileSavePath):

   
    desiredArray = parser.createDesiredArray(rawDataArray, startPoint, endPoint)
    dataOnlyArray = parser.createDataOnlyArray(desiredArray, linesUntilDataStart)


    book = load_workbook(fileSavePath)        
    writer = pd.ExcelWriter(fileSavePath, engine='openpyxl')
    writer.book = book

    if 'Power-Supply' in book.sheetnames:
        book.remove(book['Power-Supply'])


    df = pd.DataFrame(dataOnlyArray)
    df.to_excel(writer, sheet_name='Power-Supply', index=False, header=False)
    writer.save()
    


    return dataOnlyArray
